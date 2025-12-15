import tkinter as tk
from tkinter import ttk, filedialog, messagebox
# pyinstallar libreria para crear el ejecutable
from datetime import datetime ##opcion 1 de fecha
from tkcalendar import DateEntry #opcion 2 de fecha -->importar calendario parra fechas y horas
##import openpyxl si ya instalada
import openpyxl as op

def seleccionarExcel(): #ventanita de texto
    archivo = filedialog.askopenfilename(
                                         title="Seleccionar un archivo tipo Excel",
                                         filetypes=[("Archivo tipos Excel",".xlsx"),
                                                    ("Archivos tipo Python",".py")]
                                        );
    return archivo
def datosCorrectos():
   #retirnar verdadora o falso si los datos on correctos
   #por ejemplo si los impuestos son numeros, 
   #Que el cuit sea el formato correcto XX-XXXXXXXX-X por ejemplo
   return True

def guardarDatos(): #accion de guardar
   ruta = seleccionarExcel() #seleccionamos el archivo
   wb = op.load_workbook(ruta) #abrimos el archivo
   ws = wb.active #seleccionamos la primera hoja

   datosFacturacion = [tipo_factura_valor.get().strip(),
                       fecha_calendario.get().strip(),
                       factura_nro.get(),
                       CUIL_valor.get().strip(),#XX-XXXXXXXX-X
                       nombreFantacia_valor.get().strip(),
                       importeTotal.get().strip()
                       ] ##aca se guardan los datos
   if(datosCorrectos()):
      ultimaFila = ws.max_row +1 #con ws.max_row selecciono la ultima fila, con +1 paso a la siguiente
      #Guardar datos en ultima fila en columnas por cada dato
      for columaSelecionada, valorSelecionada in enumerate(datosFacturacion, start= 1): #con start arranca en 1 la columna
         ws.cell(row=ultimaFila,column=columaSelecionada,value = valorSelecionada)
      
      wb.save(ruta) #guardar en la ruta especificada
   else:
      print("Error") #usar messagebox para mostrar el error.


root =tk.Tk()
root.title("Registro Facturacion") # title para poner titulo
root.resizable(False,False) #bloquear tamaño

#===========================================#
#             Ventanas
#===========================================#
panelDerecho = tk.Frame(root, bg="#6bf67b",width=200,height=300)
panelIzquierdo = tk.Frame(root,bg="#96ee4e",width=400,height=300)
panelIzquierdo.grid(row=0,column=0)
panelDerecho.grid(row=0,column=1)

#===========================================#
#             Tipo de Factura
#===========================================#
tipo_factura_valor = tk.StringVar(value="A") #por defecto
ttk.Label(panelIzquierdo, text="Tipo de Factura:").place(x=50, y=30)
rBA= ttk.Radiobutton(panelIzquierdo, text="a", value="A", variable=tipo_factura_valor).place(x=50, y= 50)
rBB= ttk.Radiobutton(panelIzquierdo, text="b", value="B", variable=tipo_factura_valor).place(x=100, y = 50)
rBC= ttk.Radiobutton(panelIzquierdo, text="c", value="C", variable=tipo_factura_valor).place(x=150,y=50)
rBX= ttk.Radiobutton(panelIzquierdo, text ="x", value="X", variable=tipo_factura_valor).place(x=200,y=50)

#===========================================#
#             Fecha
#===========================================#
ttk.Label(panelDerecho, text=" Fecha: ").place(x=50, y= 70)
#opcion 1
#fecha_calendario = ttk.Entry(root)
#fecha_calendario.insert(0, datetime.today().strftime("%d-%m-%Y"))

#opcion 2
fecha_calendario = DateEntry(panelDerecho, date_pattern="dd-mm-yy")
##fecha_calendario.set_date(datetime.today())

fecha_calendario.place(x=50,y=90)

#===========================================#
#             Numero Factura
#===========================================#
factura_nro = tk.StringVar() ##guardar valor
ttk.Label(panelDerecho, text="Numero de factura").place(x=50, y = 110)
ttk.Entry(panelDerecho,textvariable= factura_nro).place(x=50, y= 130)

#===========================================#
#             CUIL
#===========================================#
CUIL_valor = tk.StringVar() ##guardar valor
ttk.Label(panelDerecho, text="CUIL").place(x=50, y = 150)
ttk.Entry(panelDerecho,textvariable= CUIL_valor).place(x=50, y= 170)


#===========================================#
#             NombreFantacia
#===========================================#
nombreFantacia_valor = tk.StringVar() ##guardar valor
ttk.Label(panelIzquierdo, text="Nombre Fantacia").place(x=50, y = 70)
ttk.Entry(panelIzquierdo, textvariable=nombreFantacia_valor).place(x=50, y= 90)


#===========================================#
#             Importe Total
#===========================================#
importeTotal = tk.StringVar() ##guardar valor
ttk.Label(panelIzquierdo, text="ImporteTotal").place(x=50, y = 110)
ttk.Entry(panelIzquierdo, textvariable=importeTotal).place(x=50, y= 130)


#===========================================#
#              Boton Guardar
#===========================================#
botonGuardar = tk.Button(panelIzquierdo,text="GuardarAchivo",command=guardarDatos)
botonGuardar.place(x=250, y=250)
root.mainloop()