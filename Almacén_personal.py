import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from datetime import datetime ##opcion 1 de fecha
from tkcalendar import DateEntry #opcion 2 de fecha -->importar calendario parra fechas y horas
##import openpyxl si ya instalada
import openpyxl as op

def seleccionarExcel(): #ventanita de texto
    archivo = filedialog.askopenfilename(
                                         title="Seleccionar un archivo tipo Excel",
                                         filetypes=[("Archivo tipos Excel",".xlsx")]
                                        );
    return archivo

def guardarDatos(): #accion de guardar
   ruta = seleccionarExcel() #seleccionamos el archivo
   wb = op.load_workbook(ruta) #abrimos el archivo
   ws = wb.active #seleccionamos la primera hoja

   datosFacturacion = ["A","20-11-24","30-4567577899-22"] ##aca se guardan los datos
   
   ultimaFila = ws.max_row +1 #con ws.max_row selecciono la ultima fila, con +1 paso a la siguiente
   #Guardar datos en ultima fila en columnas por cada dato
   for columaSelecionada, valorSelecionada in enumerate(datosFacturacion, start= 1): #con start arranca en 1 la columna
      ws.cell(row=ultimaFila,column=columaSelecionada,value = valorSelecionada)
   
   wb.save(ruta) #guardar en la ruta especificada



root =tk.Tk()

root.title("Registro Facturacion") # title para poner titulo
root.resizable(False,False) #bloquear tamaño
root.geometry("800x600") #tamaño fijo


#===========================================#
#             Tipo de Factura
#===========================================#

ttk.Label(root, text="Tipo de Factura:").place(x=50, y=30)
rBA= ttk.Radiobutton(root, text="A", value="A").place(x=50, y= 50)
rBB= ttk.Radiobutton(root, text="B", value="B").place(x=100, y = 50)
rBC= ttk.Radiobutton(root, text="C", value="C").place(x=150,y=50)
rBX= ttk.Radiobutton(root, text ="X", value="X").place(x=200,y=50)

#===========================================#
#             Fecha
#===========================================#

ttk.Label(root, text=" Fecha: ").place(x=50, y= 70)
#opcion 1
#fecha_calendario = ttk.Entry(root)
#fecha_calendario.insert(0, datetime.today().strftime("%d-%m-%Y"))

#opcion 2
fecha_calendario = DateEntry(root, date_pattern="dd-mm-yy")
##fecha_calendario.set_date(datetime.today())

fecha_calendario.place(x=50,y=90)

#===========================================#
#             Numero Factura
#===========================================#
factura_nro = tk.StringVar() ##guardar valor
ttk.Label(root, text="Numero de factura").place(x=50, y = 110)
ttk.Entry(root,textvariable= factura_nro).place(x=50, y= 130)

#===========================================#
#             Numero Factura
#===========================================#
botonGuardar = tk.Button(root,text="GuardarAchivo",command=guardarDatos)
botonGuardar.place(x=750, y=550)
root.mainloop()